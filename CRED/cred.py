import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.utils import column_index_from_string

def atualizar_planilha(tabela_base_path, tabela_modelo_path, banco, convenio, output_path):
    """
    Processa a tabela base (CSV) e preenche a tabela modelo (Excel),
    salvando o resultado no caminho output_path.
    
    Parâmetros:
    - tabela_base_path: caminho do CSV base
    - tabela_modelo_path: caminho do arquivo Excel modelo
    - banco: string com o nome do banco
    - convenio: string com o nome do convênio
    - output_path: caminho onde salvar o Excel preenchido
    """
    
    tabela_base = pd.read_csv(tabela_base_path, sep=';', on_bad_lines='skip')

    wb_modelo = openpyxl.load_workbook(tabela_modelo_path)
    ws_modelo = wb_modelo.active

    data_atual = datetime.now().strftime("%d/%m/%Y")

    for i, row in tabela_base.iterrows():
        try:
            operacao = row[0]
            parcelas = row[1]
            incide_sobre = row[7]
            comissao_total = row[8]
            vigencia = row[9]

            if pd.isna(vigencia):
                vigencia = data_atual

            if "até" in str(parcelas):
                parcela_inicial, parcela_final = parcelas.split(" até ")
            else:
                parcela_inicial = parcelas
                parcela_final = parcelas

            # Convertendo para string
            operacao = str(operacao)
            parcela_inicial = str(parcela_inicial)
            parcela_final = str(parcela_final)
            vigencia = str(vigencia)
            incide_sobre = str(incide_sobre)
            comissao_total = str(comissao_total)

            linha = i + 2  # linha no Excel

            ws_modelo.cell(row=linha, column=column_index_from_string('E'), value=operacao)
            ws_modelo.cell(row=linha, column=column_index_from_string('J'), value=parcela_inicial)
            ws_modelo.cell(row=linha, column=column_index_from_string('K'), value=parcela_final)
            ws_modelo.cell(row=linha, column=column_index_from_string('H'), value=vigencia)
            ws_modelo.cell(row=linha, column=column_index_from_string('U'), value=incide_sobre)
            ws_modelo.cell(row=linha, column=column_index_from_string('X'), value=comissao_total)

            ws_modelo.cell(row=linha, column=column_index_from_string('C'), value=banco)
            ws_modelo.cell(row=linha, column=column_index_from_string('D'), value=convenio)

            ws_modelo.cell(row=linha, column=column_index_from_string('Q'), value=18)
            ws_modelo.cell(row=linha, column=column_index_from_string('R'), value=999)
            ws_modelo.cell(row=linha, column=column_index_from_string('S'), value=0)
            ws_modelo.cell(row=linha, column=column_index_from_string('T'), value=100000)
            ws_modelo.cell(row=linha, column=column_index_from_string('V'), value=0)
            ws_modelo.cell(row=linha, column=column_index_from_string('W'), value=9)

            operacao_lower = operacao.lower()

            if "ativacao" in operacao_lower:
                ws_modelo.cell(row=linha, column=column_index_from_string('L'), value="PLÁSTICO")
            elif "cartao" in operacao_lower:
                ws_modelo.cell(row=linha, column=column_index_from_string('L'), value="CARTÃO")
            elif "novo" in operacao_lower:
                ws_modelo.cell(row=linha, column=column_index_from_string('L'), value="NOVO")
            elif "refin" in operacao_lower:
                if "port" in operacao_lower:
                    ws_modelo.cell(row=linha, column=column_index_from_string('L'), value="REFIN DE PORTABILIDADE")
                else:
                    ws_modelo.cell(row=linha, column=column_index_from_string('L'), value="REFIN")
            elif "port" in operacao_lower or "portabilidade" in operacao_lower:
                ws_modelo.cell(row=linha, column=column_index_from_string('L'), value="PORTABILIDADE")
            elif "ativacao" in operacao_lower and "cartao" in operacao_lower:
                ws_modelo.cell(row=linha, column=column_index_from_string('L'), value="PLÁSTICO")

            if "web plus" in operacao_lower:
                ws_modelo.cell(row=linha, column=column_index_from_string('M'), value="FÍSICO")
            else:
                ws_modelo.cell(row=linha, column=column_index_from_string('M'), value="DIGITAL")

            # Preenche células vazias com zero, exceto colunas F e I
            for col in range(1, ws_modelo.max_column + 1):
                if col in (column_index_from_string('F'), column_index_from_string('I')):
                    continue
                if ws_modelo.cell(row=linha, column=col).value is None:
                    ws_modelo.cell(row=linha, column=col, value=0)

        except IndexError:
            print(f"Erro ao processar a linha {linha}. A linha tem menos colunas que o esperado.")

    wb_modelo.save(output_path)
    return output_path
